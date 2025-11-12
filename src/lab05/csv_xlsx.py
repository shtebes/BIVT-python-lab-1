import csv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from pathlib import Path

def csv_to_xlsx(csv_path: str, xlsx_path: str) -> None:
    """
    Конвертирует CSV в XLSX.
    Использовать openpyxl ИЛИ xlsxwriter.
    Первая строка CSV — заголовок.
    Лист называется "Sheet1".
    Колонки — автоширина по длине текста (не менее 8 символов).
    Проверка ошибок:
       - неверный тип файла, пустой JSON или CSV → ValueError.
       - осутствующий файл → FileNotFoundError
    """
    wb=Workbook()
    ws=wb.active
    ws.title="Sheet1"

    csv_file=Path(csv_path)
    if not csv_file.exists():
        raise FileNotFoundError() # Файл не найден
    if csv_file.suffix != '.csv':
        raise ValueError() # Неверный тип файла
    
    # Чтение CSV данных
    with open(csv_path, 'r', encoding='utf-8') as f:
        reader= csv.DictReader(f)
        rows = list(reader)
    if len(rows)==0:
        raise ValueError("Файл не содержит данных")
    if not reader.fieldnames:
        raise ValueError("Файл не содержит заголовка")
    
    ws.append(reader.fieldnames) # Запись заголовков

    r_count=0
    for row in rows:
        r_count+=1

        data_for_ex=[]
        for title in reader.fieldnames:
            data_for_ex.append(row[title])
        ws.append(data_for_ex)
    if r_count == 0:
        raise ValueError("Нет данных")


    for col_index in range(1,len(reader.fieldnames)+1):
        column_letter=get_column_letter(col_index)
        max_len=0


        for row in ws[column_letter]:
            if row.value is not None:
                max_len=max(max_len,len(str(row.value)))

        m_width=max(max_len+2, 8)
        ws.column_dimensions[column_letter].width =m_width


    xlsx_path = Path(xlsx_path)
    
    wb.save(xlsx_path)

csv_to_xlsx("data/lab05/samples/people.csv", "data/lab05/out/people.xlsx")
csv_to_xlsx("data/lab05/samples/cities.csv", "data/lab05/out/cities.xlsx")
